# Importaciones estándar de Python
from datetime import date, datetime  # Manejo de fechas y horas

# Importaciones de Django
from django.utils.decorators import method_decorator
from django.contrib import messages  # Envío de mensajes al contexto (ejemplo: mensajes de éxito o error)
from django.contrib.auth.decorators import login_required  # Decorador para restringir acceso a usuarios autenticados
from django.contrib.auth.mixins import LoginRequiredMixin  # Mixin para vistas basadas en clases que requieren autenticación
from django.contrib.auth.models import User  # Modelo de usuarios de Django
from django.core.paginator import Paginator  # Paginación de listas de objetos
from django.db import IntegrityError  # Manejo de errores de integridad en la base de datos
from django.db.models import Q, Count, Avg  # Operadores para consultas avanzadas a la base de datos
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse  # Respuestas HTTP y JSON
from django.shortcuts import render, redirect, get_object_or_404  # Métodos para renderizar vistas y manejar redirecciones
from django.urls import reverse_lazy  # Generación de URLs reversas para redirección
from django.utils.timezone import now, timedelta  # Fechas y tiempos con soporte de zona horaria
from django.views.generic.edit import CreateView, UpdateView  # Vistas genéricas para creación y edición de objetos
# Librerías de terceros
import openpyxl  # Librería para trabajar con archivos Excel
from openpyxl.utils import get_column_letter  # Utilidad para obtener letras de columnas en Excel
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font  # Estilos y formato para celdas en Excel
from openpyxl.drawing.image import Image  # Insertar imágenes en hojas de cálculo Excel

# Framework Django Rest Framework
from rest_framework.response import Response  # Respuestas de APIs
from rest_framework.views import APIView  # Clase base para construir APIs

# Importaciones específicas del proyecto
from .forms import RegistroDeArchivoForm, FUIDForm, FichaPacienteForm  # Formularios personalizados
from .models import (  # Modelos de la base de datos
    RegistroDeArchivo,
    SubserieDocumental,
    SerieDocumental,
    FUID,
    FichaPaciente
)


@login_required
def cargar_series(request):
    series = SerieDocumental.objects.all().values('codigo', 'nombre')
    return JsonResponse(list(series), safe=False)
@login_required
def cargar_subseries(request):
    serie_id = request.GET.get('serie_id')  # esto será el id (entero)
    subseries = SubserieDocumental.objects.filter(serie_id=serie_id).values('id', 'nombre')
    return JsonResponse(list(subseries), safe=False)

from guardian.shortcuts import assign_perm  # <-- Importamos assign_perm

@login_required
# Listar registros
def lista_registros(request):
    registros = RegistroDeArchivo.objects.all()
    return render(request, 'registro_list.html', {'registros': registros})

@login_required
def crear_registro(request):
    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.creado_por = request.user  # Asigna el usuario autenticado
            registro.save()

            # 1) Asignar los permisos a nivel de objeto al usuario creador.
            assign_perm('documentos.view_own_registro', request.user, registro)
            assign_perm('documentos.edit_own_registro', request.user, registro)
            assign_perm('documentos.delete_own_registro', request.user, registro)

            # 2) Mensajes de éxito
            messages.success(request, 'Registro de archivo creado exitosamente.')

            # 3) Limpiamos el formulario para que quede listo para un nuevo registro
            form = RegistroDeArchivoForm()
        else:
            # Agrega mensajes de error para cada campo inválido
            for field, errors in form.errors.items():
                field_name = form.fields[field].label  # Obtiene la etiqueta del campo
                for error in errors:
                    messages.error(request, f"{field_name}: {error}")

    else:
        form = RegistroDeArchivoForm()
        # Subseries vacío por defecto (si no se selecciona serie)
        form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.none()

    return render(request, 'registro_form.html', {'form': form})


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden
from guardian.shortcuts import get_perms
from .models import RegistroDeArchivo, SubserieDocumental
from .forms import RegistroDeArchivoForm

@login_required
def editar_registro(request, pk):
    registro = get_object_or_404(RegistroDeArchivo, id=pk)

    # 1) Si no es superusuario, verifica que el registro sea suyo
    if not request.user.is_superuser:
        if registro.creado_por != request.user:
            return HttpResponseForbidden("No puedes editar un registro que no creaste.")

        # 2) Verifica que tenga el permiso de objeto a nivel de Guardian
        perms_del_usuario = get_perms(request.user, registro)
        if 'documentos.edit_own_registro' not in perms_del_usuario:
            return HttpResponseForbidden("No tienes permiso para editar este registro.")

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST, instance=registro)
        # Lógica de subseries
        codigo_serie = request.POST.get('codigo_serie')
        if codigo_serie:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.filter(serie_id=codigo_serie)

        if form.is_valid():
            form.save()
            return redirect('lista_registros')
    else:
        form = RegistroDeArchivoForm(instance=registro)
        if registro.codigo_serie:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.filter(serie=registro.codigo_serie)
        else:
            form.fields['codigo_subserie'].queryset = SubserieDocumental.objects.none()

    return render(request, 'registro_form.html', {'form': form})



from guardian.utils import get_anonymous_user
from guardian.shortcuts import get_perms


from guardian.shortcuts import get_perms

@login_required
def eliminar_registro(request, pk):
    registro = get_object_or_404(RegistroDeArchivo, pk=pk)

    # Si superuser, ok
    if request.user.is_superuser:
        registro.delete()
        return redirect('lista_registros')

    # Verifica si el user tiene permiso de delete a nivel de objeto
    perms = get_perms(request.user, registro)
    if 'delete_own_registro' in perms:
        registro.delete()
        return redirect('lista_registros')
    else:
        return HttpResponseForbidden("No tienes permiso para eliminar este registro.")


@login_required
def lista_completa_registros(request):
    registros = RegistroDeArchivo.objects.all()
    return render(request, 'registro_completo.html', {'registros': registros})


@login_required
def registros_api(request):
    registros = RegistroDeArchivo.objects.all()

    # Parámetros básicos
    draw = request.GET.get("draw", 1)
    start = int(request.GET.get("start", 0))
    length = int(request.GET.get("length", 10))

    # Búsqueda por columna
    # DataTables envía columns[0][data], columns[0][search][value], etc.
    i = 0
    while True:
        col_data = request.GET.get(f'columns[{i}][data]')
        if col_data is None:
            break  # no hay más columnas en el request
        col_search_value = request.GET.get(f'columns[{i}][search][value]', '').strip()

        if col_search_value:
            if col_data == 'numero_orden':
                registros = registros.filter(numero_orden__icontains=col_search_value)
            elif col_data == 'codigo':
                registros = registros.filter(codigo__icontains=col_search_value)
            elif col_data == 'codigo_serie':
                registros = registros.filter(codigo_serie__nombre__icontains=col_search_value)
            elif col_data == 'codigo_subserie':
                registros = registros.filter(codigo_subserie__nombre__icontains=col_search_value)
            elif col_data == 'unidad_documental':
                registros = registros.filter(unidad_documental__icontains=col_search_value)
            elif col_data == 'fecha_archivo':
                # Si es un texto parcial, puedes dejarlo con icontains
                registros = registros.filter(fecha_archivo__icontains=col_search_value)
            elif col_data == 'soporte_fisico':
                # Mapeo opcional si usas "✔", "✖", "True", "False", etc.
                if col_search_value in ['✔','true','True','1','si','Sí']:
                    registros = registros.filter(soporte_fisico=True)
                elif col_search_value in ['✖','false','False','0','no','No']:
                    registros = registros.filter(soporte_fisico=False)
            elif col_data == 'soporte_electronico':
                if col_search_value in ['✔','true','True','1','si','Sí']:
                    registros = registros.filter(soporte_electronico=True)
                elif col_search_value in ['✖','false','False','0','no','No']:
                    registros = registros.filter(soporte_electronico=False)
            elif col_data == 'creado_por':
                registros = registros.filter(creado_por__username__icontains=col_search_value)
            # Añade más elif si tuvieras más campos
        i += 1

    # Total sin filtros (para recordsTotal)
    total_registros = RegistroDeArchivo.objects.count()

    # Paginación
    paginator = Paginator(registros, length)
    page_number = start // length + 1
    page = paginator.get_page(page_number)

    # Construye la data de respuesta
    data = []
    for registro in page:
        data.append({
            "numero_orden": registro.numero_orden,
            "codigo": registro.codigo,
            "codigo_serie": registro.codigo_serie.nombre if registro.codigo_serie else "",
            "codigo_subserie": registro.codigo_subserie.nombre if registro.codigo_subserie else "",
            "unidad_documental": registro.unidad_documental,
            "fecha_archivo": registro.fecha_archivo,
            "soporte_fisico": registro.soporte_fisico,
            "soporte_electronico": registro.soporte_electronico,
            "creado_por": registro.creado_por.username if registro.creado_por else "N/A",
            "id": registro.id,  # importante para los enlaces Editar/Eliminar
        })

    response = {
        "draw": int(draw),
        "recordsTotal": total_registros,
        "recordsFiltered": registros.count(),
        "data": data,
    }
    return JsonResponse(response)

@login_required
def registros_api_completo(request):
    registros = RegistroDeArchivo.objects.all()

    # Paginación y parámetros de DataTables
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))

    # Filtrar por columnas
    i = 0
    while True:
        col_data = request.GET.get(f'columns[{i}][data]')
        if col_data is None:
            break
        search_value = request.GET.get(f'columns[{i}][search][value]', '').strip()
        if search_value:
            if col_data == 'numero_orden':
                registros = registros.filter(numero_orden__icontains=search_value)
            elif col_data == 'codigo':
                registros = registros.filter(codigo__icontains=search_value)
            elif col_data == 'codigo_serie':
                registros = registros.filter(codigo_serie__nombre__icontains=search_value)
            elif col_data == 'codigo_subserie':
                registros = registros.filter(codigo_subserie__nombre__icontains=search_value)
            elif col_data == 'unidad_documental':
                registros = registros.filter(unidad_documental__icontains=search_value)
            elif col_data == 'fecha_archivo':
                registros = registros.filter(fecha_archivo__icontains=search_value)
            elif col_data == 'fecha_inicial':
                registros = registros.filter(fecha_inicial__icontains=search_value)
            elif col_data == 'fecha_final':
                registros = registros.filter(fecha_final__icontains=search_value)
            elif col_data == 'soporte_fisico':
                registros = registros.filter(soporte_fisico=search_value.lower() in ['true', '1', '✔'])
            elif col_data == 'soporte_electronico':
                registros = registros.filter(soporte_electronico=search_value.lower() in ['true', '1', '✔'])
            elif col_data == 'caja':
                registros = registros.filter(caja__icontains=search_value)
            elif col_data == 'carpeta':
                registros = registros.filter(carpeta__icontains=search_value)
            elif col_data == 'ubicacion':
                registros = registros.filter(ubicacion__icontains=search_value)
            # Agrega más filtros si es necesario
        i += 1

    # Paginación
    paginator = Paginator(registros, length)
    page_number = start // length + 1
    page = paginator.get_page(page_number)

    # Preparar datos para DataTables
    data = []
    for registro in page:
        data.append({
            "numero_orden": registro.numero_orden,
            "codigo": registro.codigo,
            "codigo_serie": registro.codigo_serie.nombre if registro.codigo_serie else "",
            "codigo_subserie": registro.codigo_subserie.nombre if registro.codigo_subserie else "",
            "unidad_documental": registro.unidad_documental,
            "fecha_archivo": registro.fecha_archivo,
            "fecha_inicial": registro.fecha_inicial,
            "fecha_final": registro.fecha_final,
            "soporte_fisico": registro.soporte_fisico,
            "soporte_electronico": registro.soporte_electronico,
            "caja": registro.caja,
            "carpeta": registro.carpeta,
            "tomo_legajo_libro": registro.tomo_legajo_libro,
            "numero_folios": registro.numero_folios,
            "tipo": registro.tipo,
            "cantidad": registro.cantidad,
            "ubicacion": registro.ubicacion,
            "cantidad_documentos_electronicos": registro.cantidad_documentos_electronicos,
            "tamano_documentos_electronicos": registro.tamano_documentos_electronicos,
            "notas": registro.notas,
            "creado_por": registro.creado_por.username if registro.creado_por else "",
            "fecha_creacion": registro.fecha_creacion,
        })

    # Respuesta JSON
    response = {
        "draw": draw,
        "recordsTotal": RegistroDeArchivo.objects.count(),
        "recordsFiltered": registros.count(),
        "data": data,
    }
    return JsonResponse(response)



####
@login_required
def registros_api_con_id(request):
    registros = RegistroDeArchivo.objects.all()

    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))

    # Filtro por columnas (idéntico a tu lógica actual)
    i = 0
    while True:
        col_data = request.GET.get(f'columns[{i}][data]')
        if col_data is None:
            break
        search_value = request.GET.get(f'columns[{i}][search][value]', '').strip()

        if search_value:
            if col_data == 'numero_orden':
                registros = registros.filter(numero_orden__icontains=search_value)
            elif col_data == 'codigo':
                registros = registros.filter(codigo__icontains=search_value)
            elif col_data == 'codigo_serie':
                registros = registros.filter(codigo_serie__nombre__icontains=search_value)
            elif col_data == 'codigo_subserie':
                registros = registros.filter(codigo_subserie__nombre__icontains=search_value)
            elif col_data == 'unidad_documental':
                registros = registros.filter(unidad_documental__icontains=search_value)
            elif col_data == 'fecha_archivo':
                registros = registros.filter(fecha_archivo__icontains=search_value)
            elif col_data == 'fecha_inicial':
                registros = registros.filter(fecha_inicial__icontains=search_value)
            elif col_data == 'fecha_final':
                registros = registros.filter(fecha_final__icontains=search_value)
            elif col_data == 'soporte_fisico':
                registros = registros.filter(soporte_fisico=search_value.lower() in ['true', '1', '✔'])
            elif col_data == 'soporte_electronico':
                registros = registros.filter(soporte_electronico=search_value.lower() in ['true', '1', '✔'])
            elif col_data == 'caja':
                registros = registros.filter(caja__icontains=search_value)
            elif col_data == 'carpeta':
                registros = registros.filter(carpeta__icontains=search_value)
            elif col_data == 'ubicacion':
                registros = registros.filter(ubicacion__icontains=search_value)
            # Agrega más si fuese necesario
        i += 1

    # Paginación
    paginator = Paginator(registros, length)
    page_number = start // length + 1
    page = paginator.get_page(page_number)

    # Construye la lista de data con id incluido
    data = []
    for registro in page:
        data.append({
            "id": registro.id,  # <--- Aquí agregamos el id
            "numero_orden": registro.numero_orden,
            "codigo": registro.codigo,
            "codigo_serie": registro.codigo_serie.nombre if registro.codigo_serie else "",
            "codigo_subserie": registro.codigo_subserie.nombre if registro.codigo_subserie else "",
            "unidad_documental": registro.unidad_documental,
            "fecha_archivo": registro.fecha_archivo,
            "fecha_inicial": registro.fecha_inicial,
            "fecha_final": registro.fecha_final,
            "soporte_fisico": registro.soporte_fisico,
            "soporte_electronico": registro.soporte_electronico,
            "caja": registro.caja,
            "carpeta": registro.carpeta,
            "tomo_legajo_libro": registro.tomo_legajo_libro,
            "numero_folios": registro.numero_folios,
            "tipo": registro.tipo,
            "cantidad": registro.cantidad,
            "ubicacion": registro.ubicacion,
            "cantidad_documentos_electronicos": registro.cantidad_documentos_electronicos,
            "tamano_documentos_electronicos": registro.tamano_documentos_electronicos,
            "notas": registro.notas,
            "creado_por": registro.creado_por.username if registro.creado_por else "",
            "fecha_creacion": registro.fecha_creacion,
        })

    response = {
        "draw": draw,
        "recordsTotal": RegistroDeArchivo.objects.count(),
        "recordsFiltered": registros.count(),
        "data": data,
    }
    return JsonResponse(response)






# Vista para crear un FUID

# vistas.py (parte)

from django.template.loader import render_to_string
from django.http import JsonResponse

@login_required
def form_registro_fuid_ajax(request, fuid_id):
    """
    Retorna el HTML parcial de un formulario para crear un RegistroDeArchivo
    asociado a un FUID, listo para inyectar en un modal.
    """
    fuid = get_object_or_404(FUID, pk=fuid_id)
    form = RegistroDeArchivoForm()  # Form vacío
    
    # Renderizamos un template parcial con el formulario
    html_form = render_to_string(
        'partials/_form_registro.html',
        {'form': form, 'fuid': fuid},
        request=request
    )
    return JsonResponse({'html_form': html_form})

@login_required
def crear_registro_fuid_ajax(request, fuid_id):
    fuid = get_object_or_404(FUID, pk=fuid_id)

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.creado_por = request.user
            registro.save()

            # Asignar permisos a nivel de objeto (si usas Guardian)
            assign_perm('documentos.view_own_registro', request.user, registro)
            assign_perm('documentos.edit_own_registro', request.user, registro)
            assign_perm('documentos.delete_own_registro', request.user, registro)

            # Asociar con FUID
            fuid.registros.add(registro)

            # Devolvemos éxito y el registro en JSON, para actualizar la tabla sin recargar
            return JsonResponse({
                'ok': True,
                'message': 'Registro creado exitosamente.',
                'registro': {
                    'id': registro.id,
                    'numero_orden': registro.numero_orden,
                    'codigo': registro.codigo or '',
                    # ... agrega las demás propiedades que necesites en el JS
                }
            })
        else:
            # Devolvemos el HTML del form con errores
            html_form = render_to_string(
                'partials/_form_registro.html',
                {'form': form, 'fuid': fuid},
                request=request
            )
            return JsonResponse({'ok': False, 'html_form': html_form})

    # Si no es POST, retornamos un 405 o algo similar
    return JsonResponse({'ok': False, 'message': 'Método no permitido'}, status=405)


class FUIDCreateView(LoginRequiredMixin, CreateView):
    model = FUID
    form_class = FUIDForm
    template_name = "fuid_form.html"
    success_url = reverse_lazy("lista_fuids")

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)

        # Obtener filtros de la solicitud
        usuario = self.request.GET.get("usuario")
        fecha_inicio = self.request.GET.get("fecha_inicio")
        fecha_fin = self.request.GET.get("fecha_fin")

        # Construir queryset dinámico
        registros = RegistroDeArchivo.objects.filter(fuids__isnull=True)

        if usuario:
            registros = registros.filter(creado_por_id=usuario)
        if fecha_inicio:
            registros = registros.filter(fecha_creacion__gte=fecha_inicio)
        if fecha_fin:
            registros = registros.filter(fecha_creacion__lte=fecha_fin)

        # Ajustar el queryset de 'registros' en el formulario
        form.fields['registros'].queryset = registros
        return form

    # Mantén el decorador @login_required si lo deseas, aunque 
    # LoginRequiredMixin normalmente ya restringe el acceso a la clase
    def form_valid(self, form):
        # Asigna automáticamente el usuario que crea el FUID
        form.instance.creado_por = self.request.user
        fuid = form.save()

        # Agrega los permisos a nivel de objeto con Guardian:
        assign_perm('documentos.view_own_fuid', self.request.user, fuid)
        assign_perm('documentos.edit_own_fuid', self.request.user, fuid)
        assign_perm('documentos.delete_own_fuid', self.request.user, fuid)

        # Mantén tu lógica para asociar registros
        registros = form.cleaned_data["registros"]
        fuid.registros.set(registros)

        return super().form_valid(form)




class FUIDUpdateView(UpdateView):
    model = FUID
    form_class = FUIDForm
    template_name = "fuid_form.html"
    success_url = reverse_lazy("lista_fuids")

    def get_form_kwargs(self):
        # Pasa argumentos adicionales al formulario
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user  # Usuario autenticado
        return kwargs

    def form_valid(self, form):
        # Asigna los registros seleccionados al FUID
        fuid = form.save()
        registros = form.cleaned_data.get("registros")
        fuid.registros.set(registros)
        return super().form_valid(form)
    
@login_required
def lista_fuids(request):
    fuids = FUID.objects.all()  # Obtén todos los FUIDs
    return render(request, 'fuid_list.html', {'fuids': fuids})

@login_required
def detalle_fuid(request, pk):
    fuid = get_object_or_404(FUID, pk=pk)
    registros = fuid.registros.all()
    return render(request, 'fuid_complete_list.html', {'fuid': fuid, 'registros': registros})

@login_required
def agregar_registro_a_fuid(request, fuid_id):
    fuid = get_object_or_404(FUID, pk=fuid_id)

    if request.method == 'POST':
        form = RegistroDeArchivoForm(request.POST)
        if form.is_valid():
            # 1. Crear el registro y asignar creador
            registro = form.save(commit=False)
            registro.creado_por = request.user
            registro.save()

            # 2. (Opcional) Asignar permisos a nivel de objeto con Guardian
            assign_perm('documentos.view_own_registro', request.user, registro)
            assign_perm('documentos.edit_own_registro', request.user, registro)
            assign_perm('documentos.delete_own_registro', request.user, registro)

            # 3. Asociar el registro recién creado con el FUID
            fuid.registros.add(registro)

            messages.success(request, 'Registro creado y asociado correctamente al FUID.')
            return redirect('detalle_fuid', pk=fuid.id)  # Volver a la página detalle
        else:
            for field, errors in form.errors.items():
                field_name = form.fields[field].label
                for error in errors:
                    messages.error(request, f"{field_name}: {error}")
    else:
        form = RegistroDeArchivoForm()

    return render(request, 'agregar_registro_a_fuid.html', {
        'form': form,
        'fuid': fuid
    })

@login_required
def welcome_view(request):
    return render(request, 'welcome.html')

from django.http import HttpResponseForbidden

@login_required
def crear_ficha_paciente(request):
    # Verificar si el usuario tiene permiso global para agregar fichas
    if not request.user.has_perm('documentos.add_fichapaciente'):
        return HttpResponseForbidden("No tienes permiso para crear fichas de pacientes.")

    if request.method == 'POST':
        form = FichaPacienteForm(request.POST)
        if form.is_valid():
            ficha = form.save()
            messages.success(request, 'Ficha del paciente registrada exitosamente.')
            return redirect('crear_ficha')  # Redirige a la misma página o a otra URL
        else:
            # Manejo de errores en el formulario
            for field, errors in form.errors.items():
                field_name = form.fields[field].label
                for error in errors:
                    messages.error(request, f"{field_name}: {error}")
    else:
        form = FichaPacienteForm()

    return render(request, 'ficha_paciente_form.html', {'form': form})



@login_required
def lista_fichas_paciente(request):
    fichas = FichaPaciente.objects.all()
    return render(request, 'lista_fichas_paciente.html', {'fichas': fichas})

@method_decorator(login_required, name='dispatch')
class EditarFichaPaciente(UpdateView):
    model = FichaPaciente
    fields = '__all__'
    template_name = 'ficha_paciente_form.html'
    success_url = reverse_lazy('lista_fichas')
    pk_url_kwarg = 'consecutivo'  # Usar 'consecutivo' en lugar de 'pk'

    def dispatch(self, request, *args, **kwargs):
        # Verificar si el usuario tiene permiso global para editar fichas
        if not request.user.has_perm('documentos.change_fichapaciente'):
            return HttpResponseForbidden("No tienes permiso para editar fichas de pacientes.")
        
        # Verificar que la ficha existe
        self.object = get_object_or_404(FichaPaciente, consecutivo=kwargs.get(self.pk_url_kwarg))
        return super().dispatch(request, *args, **kwargs)



@login_required
def detalle_ficha_paciente(request, consecutivo):
    ficha = get_object_or_404(FichaPaciente, consecutivo=consecutivo)
    return render(request, 'detalle_ficha_paciente.html', {'ficha': ficha})




class ListaFichasAPIView(APIView):
    def get(self, request):
        # Parámetros enviados desde el frontend
        fecha_inicio = request.GET.get('fecha_inicio', None)
        fecha_fin = request.GET.get('fecha_fin', None)
        filtro_identificacion = request.GET.get('filtro_identificacion', None)
        filtro_historia = request.GET.get('filtro_historia', None)
        filtro_nombre = request.GET.get('filtro_nombre', None)
        filtro_similar = request.GET.get('filtro_similar', None)
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 250))

        # Ordenamiento
        order_column = int(request.GET.get('order[0][column]', 0))
        order_dir = request.GET.get('order[0][dir]', 'asc')

        # Mapear columnas de DataTables a campos del modelo
        column_mapping = {
            0: 'consecutivo',
            1: 'primer_nombre',  # Ordenar por primer nombre
            2: 'tipo_identificacion',
            3: 'num_identificacion',
            4: 'sexo',
            5: 'activo',  # Ordenar por estado
            6: 'fecha_nacimiento',
            7: 'Numero_historia_clinica',
        }

        # Determinar el campo para ordenar
        order_field = column_mapping.get(order_column, 'consecutivo')  # Campo predeterminado: consecutivo
        if order_dir == 'desc':
            order_field = f"-{order_field}"  # Prefijo "-" para orden descendente

        # Base queryset
        queryset = FichaPaciente.objects.all()

        # Filtros avanzados
        if fecha_inicio and fecha_fin:
            queryset = queryset.filter(fecha_nacimiento__range=[fecha_inicio, fecha_fin])
        if filtro_identificacion:
            queryset = queryset.filter(num_identificacion__icontains=filtro_identificacion)
        if filtro_historia:
            queryset = queryset.filter(Numero_historia_clinica__icontains=filtro_historia)
        if filtro_nombre:
            queryset = queryset.filter(
                primer_nombre__icontains=filtro_nombre
            ) | queryset.filter(
                primer_apellido__icontains=filtro_nombre
            )
        if filtro_similar:
            queryset = queryset.filter(
                primer_nombre__icontains=filtro_similar
            ) | queryset.filter(
                segundo_nombre__icontains=filtro_similar
            ) | queryset.filter(
                primer_apellido__icontains=filtro_similar
            ) | queryset.filter(
                segundo_apellido__icontains=filtro_similar
            )

        # Aplicar ordenamiento dinámico
        queryset = queryset.order_by(order_field)

        # Paginación
        total_records = queryset.count()
        paginator = Paginator(queryset, length)
        fichas = paginator.get_page(start // length + 1).object_list

        # Formato JSON para DataTables
        data = [
            {
                "consecutivo": ficha.consecutivo,
                "nombre_completo": f"{ficha.primer_nombre} {ficha.segundo_nombre or ''} {ficha.primer_apellido} {ficha.segundo_apellido}",
                "tipo_identificacion": ficha.tipo_identificacion,
                "num_identificacion": ficha.num_identificacion,
                "sexo": ficha.sexo,
                "estado": ficha.activo,
                "fecha_nacimiento": ficha.fecha_nacimiento.strftime("%Y-%m-%d"),
                "numero_historia_clinica": ficha.Numero_historia_clinica,
            }
            for ficha in fichas
        ]

        return Response(
            {
                "draw": request.GET.get("draw", 1),
                "recordsTotal": total_records,
                "recordsFiltered": total_records,
                "data": data,
            }
        )


def export_fuid_to_excel(request, pk):
    # Obtener el FUID específico
    fuid = FUID.objects.get(pk=pk)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"FUID #{fuid.id}"

    # Función para truncar valores largos
    def truncate_value(value, max_length=30):
        if not value:
            return "N/A"
        value = str(value)
        return value if len(value) <= max_length else value[:max_length - 3] + "..."

    # Crear estilos
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    header_fill = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")

    # Combinar celdas para la imagen
    ws.merge_cells(start_row=1, start_column=1, end_row=6, end_column=22)

    # Insertar la imagen
    img_path = r"D:\descargas d\xtz\pino-d-angio-c92c3fc03f2f716d1835fcf5b169efc11833deab\hospital_document_management\documentos\templates\images\fuid_logo.png"
    img = Image(img_path)
    img.width = 1000
    img.height = 120
    ws.add_image(img, "A1")

    # Mover el cursor de escritura a la fila 7 para continuar con el contenido
    current_row = 7

    # Encabezados de datos generales
    ws.cell(row=current_row, column=1, value="Campo")
    ws.cell(row=current_row, column=2, value="Valor")
    ws.cell(row=current_row, column=17, value="AÑO")
    ws.cell(row=current_row, column=18, value="MES")
    ws.cell(row=current_row, column=19, value="DÍA")
    ws.cell(row=current_row, column=20, value="N.T.")
    current_row += 1

    # Datos generales del FUID
    fuid_data = [
        ("Entidad Productora", fuid.entidad_productora.nombre if fuid.entidad_productora else "N/A", fuid.fecha_creacion.year, fuid.fecha_creacion.month, fuid.fecha_creacion.day, ""),
        ("Unidad Administrativa", fuid.unidad_administrativa.nombre if fuid.unidad_administrativa else "N/A", "", "", "", ""),
        ("Oficina Productora", fuid.oficina_productora.nombre if fuid.oficina_productora else "N/A", "", "", "", ""),
        ("Objeto", fuid.objeto.nombre if fuid.objeto else "N/A", "", "", "", ""),
    ]
    for row_data in fuid_data:
        ws.cell(row=current_row, column=1, value=row_data[0])  # Campo
        ws.cell(row=current_row, column=2, value=row_data[1])  # Valor
        ws.cell(row=current_row, column=17, value=row_data[2])  # AÑO
        ws.cell(row=current_row, column=18, value=row_data[3])  # MES
        ws.cell(row=current_row, column=19, value=row_data[4])  # DÍA
        ws.cell(row=current_row, column=20, value=row_data[5])  # N.T.
        current_row += 1

    # Aplicar bordes solo a celdas con contenido
    for row in ws.iter_rows(min_row=7, max_row=current_row-1):
        for cell in row:
            if cell.value:  # Aplica bordes solo si hay contenido
                cell.border = border

    # Espacio antes de la sección de registros
    current_row += 1
    ws.cell(row=current_row, column=1, value="")
    current_row += 1

    # Encabezados de los registros (sin "Fecha Archivo")
    headers = [
        "N° Orden", "Código", "Código Serie", "Código Subserie", "Unidad Documental",
        "Fecha Inicial", "Fecha Final", "Soporte Físico", "Soporte Electrónico",
        "Caja", "Carpeta", "Tomo/Legajo/Libro", "N° Folios", "Tipo", "Cantidad",
        "Ubicación", "Cantidad Electrónicos", "Tamaño Electrónico", "Notas", "Creado Por", "Fecha Creación"
    ]
    start_row = current_row + 1
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=start_row+3, end_column=col_idx)
        cell = ws[f"{col_letter}{start_row}"]
        cell.value = header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        if cell.value:  # Aplica bordes solo si hay contenido
            cell.border = border

    # Mover el current_row debajo de las cabeceras
    current_row = start_row + 4

    # Agregar registros asociados (sin "Fecha Archivo")
    registros = fuid.registros.all()
    if registros.exists():
        for registro in registros:
            row_data = [
                registro.numero_orden,
                truncate_value(registro.codigo or "N/A"),
                truncate_value(registro.codigo_serie.nombre if registro.codigo_serie else "N/A"),
                truncate_value(registro.codigo_subserie.nombre if registro.codigo_subserie else "N/A"),
                truncate_value(registro.unidad_documental),
                registro.fecha_inicial.strftime('%Y-%m-%d') if registro.fecha_inicial else "N/A",
                registro.fecha_final.strftime('%Y-%m-%d') if registro.fecha_final else "N/A",
                "Sí" if registro.soporte_fisico else "No",
                "Sí" if registro.soporte_electronico else "No",
                truncate_value(registro.caja or "N/A"),
                truncate_value(registro.carpeta or "N/A"),
                truncate_value(registro.tomo_legajo_libro or "N/A"),
                registro.numero_folios or "N/A",
                truncate_value(registro.tipo or "N/A"),
                registro.cantidad or "N/A",
                truncate_value(registro.ubicacion),
                registro.cantidad_documentos_electronicos or "N/A",
                truncate_value(registro.tamano_documentos_electronicos or "N/A"),
                truncate_value(registro.notas or "N/A"),
                registro.creado_por.username if registro.creado_por else "N/A",
                registro.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
            ]
            for col_idx, val in enumerate(row_data, start=1):
                c = ws.cell(row=current_row, column=col_idx, value=val)
                if c.value:  # Aplica bordes solo si hay contenido
                    c.border = border
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="Sin registros asociados")
        current_row += 1

    # Espacio antes de la sección de roles
    current_row += 1

    # Datos de roles
    roles_data = [
        ["Elaborado Por (Nombre)", truncate_value(fuid.elaborado_por_nombre or "N/A"), 
         "Entregado Por (Nombre)", truncate_value(fuid.entregado_por_nombre or "N/A"), 
         "Recibido Por (Nombre)", truncate_value(fuid.recibido_por_nombre or "N/A")],
        ["Elaborado Por (Cargo)", truncate_value(fuid.elaborado_por_cargo or "N/A"), 
         "Entregado Por (Cargo)", truncate_value(fuid.entregado_por_cargo or "N/A"), 
         "Recibido Por (Cargo)", truncate_value(fuid.recibido_por_cargo or "N/A")],
        ["Elaborado Por (Lugar)", truncate_value(fuid.elaborado_por_lugar or "N/A"), 
         "Entregado Por (Lugar)", truncate_value(fuid.entregado_por_lugar or "N/A"), 
         "Recibido Por (Lugar)", truncate_value(fuid.recibido_por_lugar or "N/A")],
        ["Firma", "", "Firma", "", "Firma", ""],
        ["Lugar", "", "Lugar", "", "Lugar", ""],
        ["Elaborado Por (Fecha)", fuid.elaborado_por_fecha.strftime('%Y-%m-%d') if fuid.elaborado_por_fecha else "N/A",
         "Entregado Por (Fecha)", fuid.entregado_por_fecha.strftime('%Y-%m-%d') if fuid.entregado_por_fecha else "N/A",
         "Recibido Por (Fecha)", fuid.recibido_por_fecha.strftime('%Y-%m-%d') if fuid.recibido_por_fecha else "N/A"],
    ]

    # Asegurar bordes para todas las celdas de roles (rango expandido)
    start_col = 1  # Columna inicial para los datos de roles
    end_col = 10  # Aumentamos el rango de columnas ocupadas
    for row_idx, row_data in enumerate(roles_data, start=current_row):
        for col_idx, val in enumerate(row_data, start=start_col):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = border  # Aplicar bordes incluso si está vacío
        current_row += 1

    # Ajustar el ancho de las columnas automáticamente
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Configurar la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=FUID_{fuid.id}.xlsx'

    wb.save(response)
    return response


# @login_required
def calcular_edad(fecha_nacimiento):
    """
    Calcula la edad actual basada en la fecha de nacimiento.
    """
    if fecha_nacimiento:
        hoy = date.today()
        return hoy.year - fecha_nacimiento.year - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
    return None
# @login_required
def estadisticas_pacientes(request):
    """
    API para devolver estadísticas de pacientes considerando varios atributos.
    """
    usuario = request.GET.get('usuario')
    pacientes = FichaPaciente.objects.all()

    if usuario:
        pacientes = pacientes.filter(creado_por__username=usuario)

    # Calcular edades
    edades = [calcular_edad(p.fecha_nacimiento) for p in pacientes if p.fecha_nacimiento]

    # Clasificar por grupos de edad
    grupos_edad = {
        "0-18": sum(1 for e in edades if e <= 18),
        "19-35": sum(1 for e in edades if 19 <= e <= 35),
        "36-60": sum(1 for e in edades if 36 <= e <= 60),
        "60+": sum(1 for e in edades if e > 60)
    }

    datos = {
        'total_pacientes': pacientes.count(),
        'por_genero': list(pacientes.values('sexo').annotate(cantidad=Count('sexo'))),
        'por_tipo_identificacion': list(pacientes.values('tipo_identificacion').annotate(cantidad=Count('tipo_identificacion'))),
        'activos': pacientes.filter(activo=True).count(),
        'promedio_edad': round(sum(edades) / len(edades), 2) if edades else None,
        'grupos_edad': grupos_edad
    }

    return JsonResponse(datos, safe=False)

# @login_required
def estadisticas_registros(request):
    """
    API para devolver estadísticas de registros, organizados por series documentales y tipos.
    """
    try:
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        registros = RegistroDeArchivo.objects.all()

        # Filtrar por rango de fechas si se proporcionan
        if fecha_inicio and fecha_fin:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
            registros = registros.filter(fecha_archivo__range=(fecha_inicio, fecha_fin))

        # Generar estadísticas
        datos = {
            'total_registros': registros.count(),
            'por_serie': list(
                registros.values('codigo_serie__nombre').annotate(cantidad=Count('id'))
            ),
            'por_soporte': list(
                registros.values('soporte_fisico', 'soporte_electronico').annotate(cantidad=Count('id'))
            ),
            'por_tipo': list(
                registros.values('tipo').annotate(cantidad=Count('id'))
            ),
        }

        return JsonResponse(datos, safe=False)
    except Exception as e:
        print("Error en estadisticas_registros:", e)
        return JsonResponse({"error": str(e)}, status=500)



# @login_required
def estadisticas_fuids(request):
    """
    API para devolver estadísticas de FUIDs, organizados por oficinas productoras.
    """
    usuario = request.GET.get('usuario')
    fuids = FUID.objects.all()

    if usuario:
        fuids = fuids.filter(creado_por__username=usuario)

    datos = {
        'total_fuids': fuids.count(),
        'por_oficina': list(fuids.values('oficina_productora__nombre').annotate(cantidad=Count('id'))),
        'por_objeto': list(fuids.values('objeto__nombre').annotate(cantidad=Count('id'))),
        'por_entidad': list(fuids.values('entidad_productora__nombre').annotate(cantidad=Count('id'))),
    }

    return JsonResponse(datos, safe=False)

# @login_required
def pagina_estadisticas(request):
    """
    Página principal para mostrar gráficos de las estadísticas.
    """
    return render(request, 'pagina_estadisticas.html')


@login_required
def obtener_usuarios(request):
    usuarios = User.objects.values('username')
    return JsonResponse(list(usuarios), safe=False)

# mixins.py
from django.http import HttpResponseForbidden

class OficinaFilterMixin:
    """
    Filtra los objetos para que el usuario solo vea y manipule
    aquellos creados por su oficina. También bloquea la edición
    de objetos de otras oficinas.
    """
    def get_queryset(self):
        qs = super().get_queryset()
        # Si deseas que el superusuario vea todo, déjalo pasar:
        if self.request.user.is_superuser:
            return qs
        # Caso contrario, filtra por la oficina del perfil
        return qs.filter(oficina_productora=self.request.user.perfil.oficina)

    def dispatch(self, request, *args, **kwargs):
        # Bloqueo adicional para edición/eliminación
        if hasattr(self, 'get_object'):
            obj = self.get_object()
            if (not request.user.is_superuser) and (obj.oficina_productora != request.user.perfil.oficina):
                return HttpResponseForbidden("No tienes permiso sobre este recurso.")
        return super().dispatch(request, *args, **kwargs)



